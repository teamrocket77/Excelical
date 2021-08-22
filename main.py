import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from ics import Calendar, Event


c = Calendar()
d = datetime.datetime(2021, 1, 1)
# book names goes here
schedule = load_workbook('Schedule.xlsx')

# Me presuming that you have only 1 Worksheet
sheet_1 = schedule.active


"""
The following assumes that your worksheet is labeled
Day, Date, Chapter or subject, row_x, Quiz Date, Homework


With all of the data of course following in rows.
For me I picked out the 2nd(ie row [1])
column 2 column 5 and 6 are all of importance to me """
skip = False
for row in sheet_1.iter_rows():
    e = Event()

    Rows = {
        'Date': row[1],
        'Homework': row[5],
        'Test': row[4]
    }
    if Rows['Homework'].value is not None and type(Rows['Date'].value) is type(d):
        print(f"Date: {Rows['Date'].value} Event: {Rows['Homework'].value} ")
        e.name = Rows['Homework'].value
        e.begin = Rows['Date'].value
        c.events.add(e)

    if Rows['Test'].value is not None and type(Rows['Date'].value) is type(d):
        print(f"Date: {Rows['Date'].value} Event: {Rows['Test'].value}")
        e.name = Rows['Test'].value
        e.begin = Rows['Date'].value
        c.events.add(e)

# print(c.events)
with open('Prob.ics', 'w') as f:
    f.write(str(c))
    f.close()